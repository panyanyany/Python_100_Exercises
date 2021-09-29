from datetime import datetime, timedelta

import openpyxl

excel = openpyxl.Workbook()


def create_sheet(excel, title):
    print(f'正在创建表格：{title}')
    sheet = excel.create_sheet(title)
    sheet['A1'] = '姓名'
    sheet['A2'] = '张三'
    sheet['A3'] = '李四'
    sheet['A4'] = '王五'

    sheet['B1'] = '销售额'

    sheet['C1'] = '利润'


dt_beg = datetime(2021, 10, 1)
dt_end = datetime(2021, 11, 1)
total_day = (dt_end - dt_beg).days
for day in range(total_day):
    now = dt_beg + timedelta(days=day)
    create_sheet(excel, now.date().strftime("%Y-%m-%d"))

del excel['Sheet']
excel.save('30天表格.xlsx')
